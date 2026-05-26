"""
app.py  —  Streamlit web app version of the Xero P&L refresher.

Deploy free at streamlit.io/cloud — users open a URL in their browser,
no Python or software install required.

How it works:
  1. User uploads their Excel P&L template
  2. Clicks "Connect to Xero" — Xero login opens in the same tab
  3. After login Xero redirects back here with an auth code in the URL
  4. App exchanges the code for an access token
  5. Fetches P&L data for every month column in the spreadsheet
  6. Shows a preview table and lets the user download the updated file

Secrets needed (set in Streamlit Cloud → App settings → Secrets):
    XERO_CLIENT_ID     = "your-client-id"
    XERO_CLIENT_SECRET = "your-client-secret"

The Xero app redirect URI must be set to your Streamlit app's URL,
e.g. https://your-app-name.streamlit.app
"""

import streamlit as st
import base64, hashlib, json, os, secrets, time, urllib.parse, urllib.request
import io
from datetime import datetime, date
from calendar import monthrange
from collections import defaultdict

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

# ── Page config  ──────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="Xero P&L Refresher",
    page_icon="📊",
    layout="centered"
)

# ── Xero endpoints  ───────────────────────────────────────────────────────────
AUTH_URL    = "https://login.xero.com/identity/connect/authorize"
TOKEN_URL   = "https://identity.xero.com/connect/token"
TENANTS_URL = "https://api.xero.com/connections"
API_BASE    = "https://api.xero.com/api.xro/2.0"
SCOPES      = "accounting.reports.profitandloss.read accounting.settings.read offline_access"

NZD_FMT = '#,##0;(#,##0);"-"'

# ── Load secrets  ─────────────────────────────────────────────────────────────
# In Streamlit Cloud these come from the Secrets manager.
# For local testing, create a .streamlit/secrets.toml file:
#   XERO_CLIENT_ID     = "..."
#   XERO_CLIENT_SECRET = "..."

def get_secret(key):
    """Gets a secret from Streamlit secrets, or falls back to an env variable."""
    try:
        return st.secrets[key]
    except (KeyError, FileNotFoundError):
        return os.environ.get(key, "")

CLIENT_ID     = get_secret("XERO_CLIENT_ID")
CLIENT_SECRET = get_secret("XERO_CLIENT_SECRET")

# ── Helpers: OAuth  ───────────────────────────────────────────────────────────

def get_redirect_uri():
    """
    Works out the correct redirect URI for this environment.
    In production (Streamlit Cloud) this is the app's public URL.
    Locally it falls back to localhost.
    """
    try:
        base = st.context.url
        if base:
            parsed = urllib.parse.urlparse(base)
            return f"{parsed.scheme}://{parsed.netloc}"
    except Exception:
        pass
    return "http://localhost:8501"


def build_auth_url():
    """
    Creates the Xero authorisation URL with PKCE security.
    Stores the verifier and state in session state so we can use them
    when Xero redirects back.
    """
    verifier  = secrets.token_urlsafe(64)
    challenge = base64.urlsafe_b64encode(
        hashlib.sha256(verifier.encode()).digest()
    ).rstrip(b"=").decode()
    state = secrets.token_urlsafe(16)

    st.session_state["pkce_verifier"] = verifier
    st.session_state["oauth_state"]   = state

    params = urllib.parse.urlencode({
        "response_type":         "code",
        "client_id":             CLIENT_ID,
        "redirect_uri":          get_redirect_uri(),
        "scope":                 SCOPES,
        "state":                 state,
        "code_challenge":        challenge,
        "code_challenge_method": "S256",
    })
    return f"{AUTH_URL}?{params}"


def exchange_code_for_token(code):
    """
    Swaps the one-time auth code (from the URL after Xero login)
    for an access token we can use to call the API.
    """
    creds = base64.b64encode(f"{CLIENT_ID}:{CLIENT_SECRET}".encode()).decode()
    data  = urllib.parse.urlencode({
        "grant_type":    "authorization_code",
        "code":          code,
        "redirect_uri":  get_redirect_uri(),
        "code_verifier": st.session_state.get("pkce_verifier", ""),
    }).encode()

    req = urllib.request.Request(TOKEN_URL, data=data, method="POST")
    req.add_header("Content-Type",  "application/x-www-form-urlencoded")
    req.add_header("Authorization", f"Basic {creds}")

    with urllib.request.urlopen(req) as r:
        tok = json.loads(r.read())
    tok["obtained_at"] = int(time.time())
    return tok


def is_token_valid():
    """Checks if the stored access token is still valid (not expired)."""
    tok = st.session_state.get("xero_token")
    if not tok:
        return False
    expires = tok.get("obtained_at", 0) + tok.get("expires_in", 1800) - 60
    return time.time() < expires


# ── Helpers: Xero API  ────────────────────────────────────────────────────────

def xero_get(path, params=None):
    """Sends a GET request to the Xero API and returns the parsed JSON."""
    token  = st.session_state["xero_token"]["access_token"]
    tenant = st.session_state["xero_tenant_id"]

    url = f"{API_BASE}/{path}"
    if params:
        url += "?" + urllib.parse.urlencode(params)

    req = urllib.request.Request(url)
    req.add_header("Authorization",  f"Bearer {token}")
    req.add_header("Xero-Tenant-Id", tenant)
    req.add_header("Accept",         "application/json")

    with urllib.request.urlopen(req) as r:
        return json.loads(r.read())


def get_tenants():
    """Returns a list of Xero organisations this token can access."""
    token = st.session_state["xero_token"]["access_token"]
    req   = urllib.request.Request(TENANTS_URL)
    req.add_header("Authorization", f"Bearer {token}")
    req.add_header("Accept",        "application/json")
    with urllib.request.urlopen(req) as r:
        return json.loads(r.read())


def fetch_pnl(from_date, to_date):
    """
    Fetches the Xero P&L report for the given date range.
    Returns a dict: { "sales": 85000.0, "advertising": 3200.0, ... }
    All keys are lowercase for case-insensitive matching.
    """
    data = xero_get("Reports/ProfitAndLoss", {
        "fromDate":       from_date.strftime("%Y-%m-%d"),
        "toDate":         to_date.strftime("%Y-%m-%d"),
        "standardLayout": "true",
        "paymentsOnly":   "false",
    })
    results = {}
    rows = data.get("Reports", [{}])[0].get("Rows", [])

    def walk(rows_list):
        for row in rows_list:
            rt = row.get("RowType", "")
            if rt in ("Section", "SummaryRow"):
                walk(row.get("Rows", []))
            elif rt == "Row":
                cells = row.get("Cells", [])
                if len(cells) >= 2:
                    name = cells[0].get("Value", "").strip()
                    val  = cells[1].get("Value", "0") or "0"
                    try:
                        amount = float(val.replace(",", ""))
                    except ValueError:
                        amount = 0.0
                    if name:
                        results[name.lower()] = amount

    walk(rows)
    return results


# ── Helpers: Excel  ───────────────────────────────────────────────────────────

MONTH_MAP = {
    "jan":1,"feb":2,"mar":3,"apr":4,"may":5,"jun":6,
    "jul":7,"aug":8,"sep":9,"oct":10,"nov":11,"dec":12,
    "january":1,"february":2,"march":3,"april":4,"june":6,
    "july":7,"august":8,"september":9,"october":10,"november":11,"december":12,
}

def parse_month_header(val):
    """Converts a cell value like 'May 2026' to (2026, 5), or returns None."""
    if isinstance(val, (datetime, date)):
        return val.year, val.month
    if not isinstance(val, str):
        return None
    parts = val.strip().split()
    if len(parts) == 2:
        mon = MONTH_MAP.get(parts[0][:3].lower())
        try:
            yr = int(parts[1])
            if mon and 2000 <= yr <= 2100:
                return yr, mon
        except ValueError:
            pass
    return None


def read_excel(file_bytes):
    """
    Opens the uploaded Excel file from bytes and reads:
      - periods:  [(col, year, month), ...]   from row 5
      - accounts: [(row, name), ...]           from column A (non-bold rows only)
    """
    wb = load_workbook(io.BytesIO(file_bytes))
    ws = wb.active

    periods = []
    for col in range(2, ws.max_column + 1):
        parsed = parse_month_header(ws.cell(row=5, column=col).value)
        if parsed:
            periods.append((col, parsed[0], parsed[1]))

    accounts = []
    for row in range(6, ws.max_row + 1):
        cell = ws.cell(row=row, column=1)
        if not cell.value:
            continue
        if cell.font and cell.font.bold:
            continue
        accounts.append((row, str(cell.value).strip()))

    return wb, periods, accounts


def write_excel(wb, periods, accounts, all_data, org_name):
    """
    Writes Xero data into the yellow cells of the workbook.
    Returns (bytes, written_count, not_found_list, preview_dict).
    """
    ws = wb.active
    ws["A2"].value = org_name
    ws["A3"].value = f"Refreshed from Xero: {datetime.now().strftime('%d %b %Y %H:%M')}"

    written   = 0
    not_found = []
    preview   = {}

    for row, account_name in accounts:
        key = account_name.lower()
        preview[account_name] = {}

        for col, year, month in periods:
            month_label = date(year, month, 1).strftime("%b %Y")
            period_data = all_data.get((year, month), {})
            cell        = ws.cell(row=row, column=col)
            amount      = None

            if key in period_data:
                amount = period_data[key]
            else:
                matches = [(k, v) for k, v in period_data.items()
                           if key in k or k in key]
                if len(matches) == 1:
                    amount = matches[0][1]

            if amount is not None:
                cell.value         = amount
                cell.number_format = NZD_FMT
                cell.font          = Font(name="Arial", size=10)
                cell.fill          = PatternFill("solid", fgColor="FFFFFF")
                cell.alignment     = Alignment(horizontal="right")
                preview[account_name][month_label] = amount
                written += 1
            else:
                cell.fill  = PatternFill("solid", fgColor="FFFF00")
                cell.value = None
                preview[account_name][month_label] = None
                if account_name not in not_found:
                    not_found.append(account_name)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue(), written, not_found, preview


# ══════════════════════════════════════════════════════════════════════════════
# UI
# ══════════════════════════════════════════════════════════════════════════════

st.title("📊 Xero P&L Refresher")
st.caption("Upload your Excel template → connect to Xero → download the updated file")

# ── Guard: credentials check  ─────────────────────────────────────────────────
if not CLIENT_ID or not CLIENT_SECRET:
    st.error(
        "**Xero app credentials not configured.**\n\n"
        "Add `XERO_CLIENT_ID` and `XERO_CLIENT_SECRET` to your "
        "Streamlit secrets (Settings → Secrets in the Streamlit Cloud dashboard)."
    )
    st.stop()

# ── Handle OAuth callback  ────────────────────────────────────────────────────
params = st.query_params

if "code" in params and not is_token_valid():
    returned_state = params.get("state", "")
    expected_state = st.session_state.get("oauth_state", "")

    if returned_state != expected_state:
        st.error("Security check failed (state mismatch). Please try connecting again.")
    else:
        with st.spinner("Completing Xero login…"):
            try:
                tok = exchange_code_for_token(params["code"])
                st.session_state["xero_token"] = tok
                tenants = get_tenants()
                if tenants:
                    st.session_state["xero_tenant_id"]   = tenants[0]["tenantId"]
                    st.session_state["xero_tenant_name"] = tenants[0]["tenantName"]
                st.query_params.clear()
                st.rerun()
            except Exception as e:
                st.error(f"Login failed: {e}")

# ── Step 1: Upload file  ──────────────────────────────────────────────────────
st.subheader("Step 1 — Upload your P&L template")
uploaded = st.file_uploader(
    "Upload the Excel P&L template",
    type=["xlsx"],
    help="The Xero_PnL_Template.xlsx file (or any Excel file with the same layout)"
)

if uploaded:
    file_bytes = uploaded.read()
    try:
        wb, periods, accounts = read_excel(file_bytes)
        st.success(
            f"✅ Found **{len(periods)} month columns** and "
            f"**{len(accounts)} account rows** in your spreadsheet."
        )
        with st.expander("Preview what was found in the spreadsheet"):
            col1, col2 = st.columns(2)
            with col1:
                st.caption("Month columns")
                for _, y, m in periods:
                    st.write(f"• {date(y, m, 1).strftime('%B %Y')}")
            with col2:
                st.caption("Accounts to fetch")
                for _, name in accounts:
                    st.write(f"• {name}")
    except Exception as e:
        st.error(f"Could not read the file: {e}")
        st.stop()
else:
    st.info("👆 Upload your Excel template to get started.")
    st.stop()

st.divider()

# ── Step 2: Connect to Xero  ──────────────────────────────────────────────────
st.subheader("Step 2 — Connect to Xero")

if is_token_valid():
    org = st.session_state.get("xero_tenant_name", "your organisation")
    st.success(f"✅ Connected to Xero: **{org}**")

    if st.button("Disconnect / switch organisation"):
        for key in ["xero_token", "xero_tenant_id", "xero_tenant_name",
                    "pkce_verifier", "oauth_state"]:
            st.session_state.pop(key, None)
        st.rerun()

    try:
        tenants = get_tenants()
        if len(tenants) > 1:
            names   = [t["tenantName"] for t in tenants]
            current = st.session_state.get("xero_tenant_name", names[0])
            chosen  = st.selectbox("Organisation", names, index=names.index(current))
            if chosen != current:
                t = next(t for t in tenants if t["tenantName"] == chosen)
                st.session_state["xero_tenant_id"]   = t["tenantId"]
                st.session_state["xero_tenant_name"] = t["tenantName"]
                st.rerun()
    except Exception:
        pass

else:
    st.info("Click the button below to log in to Xero.")
    auth_url = build_auth_url()
    st.markdown(
        f'<a href="{auth_url}" target="_self" style="'
        'display:inline-block;background:#1ab394;color:white;padding:10px 24px;'
        'border-radius:6px;text-decoration:none;font-weight:500;font-size:14px">'
        '🔗 Connect to Xero</a>',
        unsafe_allow_html=True
    )
    st.stop()

st.divider()

# ── Step 3: Fetch and download  ───────────────────────────────────────────────
st.subheader("Step 3 — Fetch data and download")

if st.button("🔄 Fetch P&L data from Xero", type="primary", use_container_width=True):

    org_name = st.session_state.get("xero_tenant_name", "Xero")
    all_data = {}
    progress = st.progress(0, text="Starting…")
    status   = st.empty()

    for i, (col, year, month) in enumerate(periods):
        from_date   = date(year, month, 1)
        to_date     = date(year, month, monthrange(year, month)[1])
        month_label = from_date.strftime("%B %Y")

        progress.progress(i / len(periods), text=f"Fetching {month_label}…")
        status.caption(f"Calling Xero API for {month_label}…")

        try:
            all_data[(year, month)] = fetch_pnl(from_date, to_date)
        except Exception as e:
            st.warning(f"Could not fetch {month_label}: {e}")
            all_data[(year, month)] = {}

    progress.progress(1.0, text="Writing to spreadsheet…")

    wb, _, _ = read_excel(file_bytes)
    result_bytes, written, not_found, preview = write_excel(
        wb, periods, accounts, all_data, org_name
    )

    progress.empty()
    status.empty()

    st.success(f"✅ **{written} cells updated** across {len(periods)} months.")

    if not_found:
        with st.expander(f"⚠️ {len(not_found)} account(s) not matched — click to review"):
            st.caption("Check spelling matches Xero exactly.")
            for a in not_found:
                st.write(f"• {a}")

    if preview:
        with st.expander("Preview data", expanded=True):
            import pandas as pd
            month_labels = [date(y, m, 1).strftime("%b %Y") for _, y, m in periods]
            rows_data = []
            for acct, months in preview.items():
                row = {"Account": acct}
                for lbl in month_labels:
                    val = months.get(lbl)
                    row[lbl] = f"${val:,.0f}" if val is not None else "—"
                rows_data.append(row)
            st.dataframe(pd.DataFrame(rows_data), use_container_width=True, hide_index=True)

    filename = f"PnL_Xero_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    st.download_button(
        label="⬇️ Download updated Excel file",
        data=result_bytes,
        file_name=filename,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
        type="primary",
    )
