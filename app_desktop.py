"""
app_desktop.py  —  Desktop app version (tkinter GUI).
Build into a standalone .exe or .app with PyInstaller — no Python needed to run it.

Build commands (run once on the target machine):
    pip install pyinstaller openpyxl python-dateutil
    pyinstaller --onefile --windowed --name "Xero PnL Refresher" app_desktop.py
    
The resulting file is in the dist/ folder.
"""

import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext
import base64, hashlib, json, os, secrets, sys, threading, time
import urllib.parse, urllib.request, webbrowser
from datetime import datetime, date
from calendar import monthrange
from http.server import HTTPServer, BaseHTTPRequestHandler
from io import StringIO

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ── Xero constants  ───────────────────────────────────────────────────────────
AUTH_URL    = "https://login.xero.com/identity/connect/authorize"
TOKEN_URL   = "https://identity.xero.com/connect/token"
TENANTS_URL = "https://api.xero.com/connections"
API_BASE    = "https://api.xero.com/api.xro/2.0"
REDIRECT_URI  = "http://localhost:8765/callback"
CALLBACK_PORT = 8765
SCOPES = "accounting.reports.profitandloss.read accounting.settings.read offline_access"
TOKEN_FILE = os.path.join(os.path.expanduser("~"), ".xero_pnl_token.json")
CONFIG_FILE = os.path.join(os.path.expanduser("~"), ".xero_pnl_config.json")
NZD_FMT = '#,##0;(#,##0);"-"'

# ── OAuth callback handler  ───────────────────────────────────────────────────
class _CB(BaseHTTPRequestHandler):
    code = None; expected = None
    def do_GET(self):
        qs = urllib.parse.parse_qs(urllib.parse.urlparse(self.path).query)
        if qs.get("state", [None])[0] == _CB.expected:
            _CB.code = qs.get("code", [None])[0]
        self._send(200, "<h2 style='font-family:Arial;color:#1D7A6B'>✅ Done — return to the app.</h2>")
    def _send(self, s, b):
        self.send_response(s); self.send_header("Content-Type","text/html"); self.end_headers()
        self.wfile.write(b.encode())
    def log_message(self, *a): pass

def _post(url, data, auth=None):
    enc = urllib.parse.urlencode(data).encode()
    req = urllib.request.Request(url, data=enc, method="POST")
    req.add_header("Content-Type", "application/x-www-form-urlencoded")
    if auth: req.add_header("Authorization", f"Basic {auth}")
    with urllib.request.urlopen(req) as r:
        return json.loads(r.read())

def _b64(cid, sec):
    return base64.b64encode(f"{cid}:{sec}".encode()).decode()

# ── Excel helpers  ─────────────────────────────────────────────────────────────
MONTH_MAP = {
    "jan":1,"feb":2,"mar":3,"apr":4,"may":5,"jun":6,
    "jul":7,"aug":8,"sep":9,"oct":10,"nov":11,"dec":12,
    "january":1,"february":2,"march":3,"april":4,"june":6,
    "july":7,"august":8,"september":9,"october":10,"november":11,"december":12,
}

def parse_month(val):
    if isinstance(val, (datetime, date)): return val.year, val.month
    if not isinstance(val, str): return None
    p = val.strip().split()
    if len(p) == 2:
        m = MONTH_MAP.get(p[0][:3].lower())
        try:
            y = int(p[1])
            if m and 2000 <= y <= 2100: return y, m
        except ValueError: pass
    return None

def read_config(wb):
    ws = wb.active
    periods  = [(c, *parse_month(ws.cell(5,c).value))
                for c in range(2, ws.max_column+1)
                if parse_month(ws.cell(5,c).value)]
    accounts = [(r, str(ws.cell(r,1).value).strip())
                for r in range(6, ws.max_row+1)
                if ws.cell(r,1).value and not (ws.cell(r,1).font and ws.cell(r,1).font.bold)]
    return periods, accounts

# ══════════════════════════════════════════════════════════════════════════════
# MAIN APPLICATION CLASS
# ══════════════════════════════════════════════════════════════════════════════
class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Xero P&L Refresher")
        self.geometry("620x680")
        self.resizable(True, True)
        self.configure(bg="#f5f5f5")

        # State
        self.token       = None
        self.tenant_id   = None
        self.tenant_name = None
        self.workbook_path = tk.StringVar()
        self.client_id   = tk.StringVar()
        self.client_secret = tk.StringVar()
        self._load_config()
        self._load_token()
        self._build_ui()

    # ── Config persistence  ───────────────────────────────────────────────────
    def _load_config(self):
        if os.path.exists(CONFIG_FILE):
            try:
                c = json.load(open(CONFIG_FILE))
                self.client_id.set(c.get("client_id", ""))
                self.client_secret.set(c.get("client_secret", ""))
            except: pass

    def _save_config(self):
        with open(CONFIG_FILE, "w") as f:
            json.dump({"client_id": self.client_id.get(),
                       "client_secret": self.client_secret.get()}, f)

    def _load_token(self):
        if os.path.exists(TOKEN_FILE):
            try:
                tok = json.load(open(TOKEN_FILE))
                expires = tok.get("obtained_at",0) + tok.get("expires_in",1800) - 60
                if time.time() < expires:
                    self.token = tok
            except: pass

    def _save_token(self):
        with open(TOKEN_FILE, "w") as f:
            json.dump(self.token, f, indent=2)
        try: os.chmod(TOKEN_FILE, 0o600)
        except: pass

    # ── UI layout  ────────────────────────────────────────────────────────────
    def _build_ui(self):
        style = ttk.Style(self)
        style.theme_use("clam")
        style.configure("TLabel", background="#f5f5f5", font=("Arial", 10))
        style.configure("TEntry", font=("Arial", 10))
        style.configure("Header.TLabel", font=("Arial", 11, "bold"), foreground="#1B3A5C")
        style.configure("Green.TLabel", foreground="#0F6E56", background="#f5f5f5")
        style.configure("Red.TLabel",   foreground="#993C1D", background="#f5f5f5")
        style.configure("Primary.TButton", font=("Arial", 10, "bold"))

        pad = dict(padx=16, pady=6)

        # ── Credentials section  ─────────────────────────────────────────────
        cred_frame = ttk.LabelFrame(self, text=" Xero App Credentials ", padding=10)
        cred_frame.pack(fill="x", **pad, pady=(16,6))

        ttk.Label(cred_frame, text="Client ID:").grid(row=0, column=0, sticky="w", pady=4)
        ttk.Entry(cred_frame, textvariable=self.client_id, width=46).grid(row=0, column=1, padx=8)
        ttk.Label(cred_frame, text="Client Secret:").grid(row=1, column=0, sticky="w", pady=4)
        ttk.Entry(cred_frame, textvariable=self.client_secret, width=46, show="•").grid(row=1, column=1, padx=8)
        ttk.Label(cred_frame, text="Get these from developer.xero.com → My Apps",
                  font=("Arial", 9), foreground="#888").grid(row=2, column=0, columnspan=2, sticky="w", pady=2)
        ttk.Button(cred_frame, text="Save credentials", command=self._save_config).grid(
            row=2, column=1, sticky="e")

        # ── Connection section  ──────────────────────────────────────────────
        conn_frame = ttk.LabelFrame(self, text=" Xero Connection ", padding=10)
        conn_frame.pack(fill="x", **pad)

        self.conn_status = ttk.Label(conn_frame, text="Not connected", style="Red.TLabel")
        self.conn_status.pack(side="left", expand=True, anchor="w")

        self.conn_btn = ttk.Button(conn_frame, text="Connect to Xero",
                                   command=self._connect, style="Primary.TButton")
        self.conn_btn.pack(side="right")

        self._update_connection_label()

        # ── File section  ────────────────────────────────────────────────────
        file_frame = ttk.LabelFrame(self, text=" Excel Spreadsheet ", padding=10)
        file_frame.pack(fill="x", **pad)

        ttk.Entry(file_frame, textvariable=self.workbook_path, width=46).pack(side="left", expand=True, fill="x")
        ttk.Button(file_frame, text="Browse…", command=self._browse).pack(side="right", padx=(8,0))

        # ── Run button  ──────────────────────────────────────────────────────
        self.run_btn = ttk.Button(self, text="🔄  Fetch from Xero & Update Spreadsheet",
                                  command=self._run, style="Primary.TButton")
        self.run_btn.pack(fill="x", **pad, pady=(4,4))

        # ── Progress bar  ────────────────────────────────────────────────────
        self.progress = ttk.Progressbar(self, mode="determinate")
        self.progress.pack(fill="x", padx=16, pady=4)

        # ── Log area  ────────────────────────────────────────────────────────
        log_frame = ttk.LabelFrame(self, text=" Log ", padding=6)
        log_frame.pack(fill="both", expand=True, **pad)

        self.log = scrolledtext.ScrolledText(log_frame, height=14, font=("Courier", 9),
                                              state="disabled", bg="white", relief="flat")
        self.log.pack(fill="both", expand=True)

        self._log("Ready. Upload your Excel template and connect to Xero to get started.")

    def _log(self, msg):
        """Appends a timestamped message to the log area."""
        self.log.configure(state="normal")
        self.log.insert("end", f"[{datetime.now().strftime('%H:%M:%S')}]  {msg}\n")
        self.log.see("end")
        self.log.configure(state="disabled")

    def _update_connection_label(self):
        if self.token and self.tenant_name:
            self.conn_status.configure(
                text=f"✅  Connected: {self.tenant_name}", style="Green.TLabel")
            self.conn_btn.configure(text="Reconnect")
        else:
            self.conn_status.configure(text="Not connected", style="Red.TLabel")
            self.conn_btn.configure(text="Connect to Xero")

    # ── Browse for file  ─────────────────────────────────────────────────────
    def _browse(self):
        path = filedialog.askopenfilename(
            title="Select your Excel P&L template",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )
        if path:
            self.workbook_path.set(path)
            self._log(f"File selected: {os.path.basename(path)}")

    # ── OAuth connect  ────────────────────────────────────────────────────────
    def _connect(self):
        if not self.client_id.get() or not self.client_secret.get():
            messagebox.showerror("Missing credentials",
                "Please enter your Xero Client ID and Client Secret first.")
            return
        self._save_config()
        threading.Thread(target=self._do_oauth, daemon=True).start()

    def _do_oauth(self):
        self._log("Opening Xero login in your browser…")
        self.conn_btn.configure(state="disabled")

        verifier  = secrets.token_urlsafe(64)
        challenge = base64.urlsafe_b64encode(
            hashlib.sha256(verifier.encode()).digest()).rstrip(b"=").decode()
        state = secrets.token_urlsafe(16)
        _CB.expected = state; _CB.code = None

        params = urllib.parse.urlencode({
            "response_type":"code", "client_id":self.client_id.get(),
            "redirect_uri":REDIRECT_URI, "scope":SCOPES,
            "state":state, "code_challenge":challenge,
            "code_challenge_method":"S256"})
        webbrowser.open(f"{AUTH_URL}?{params}")

        srv = HTTPServer(("localhost", CALLBACK_PORT), _CB)
        t = threading.Thread(target=srv.handle_request, daemon=True)
        t.start(); t.join(timeout=120)

        if not _CB.code:
            self.after(0, lambda: messagebox.showerror(
                "Timeout", "No response from Xero within 2 minutes. Try again."))
            self.after(0, lambda: self.conn_btn.configure(state="normal"))
            return

        try:
            tok = _post(TOKEN_URL, {
                "grant_type":"authorization_code", "code":_CB.code,
                "redirect_uri":REDIRECT_URI, "code_verifier":verifier},
                auth=_b64(self.client_id.get(), self.client_secret.get()))
            tok["obtained_at"] = int(time.time())
            self.token = tok
            self._save_token()

            # Get tenant
            req = urllib.request.Request(TENANTS_URL)
            req.add_header("Authorization", f"Bearer {tok['access_token']}")
            req.add_header("Accept", "application/json")
            with urllib.request.urlopen(req) as r:
                tenants = json.loads(r.read())
            if tenants:
                self.tenant_id   = tenants[0]["tenantId"]
                self.tenant_name = tenants[0]["tenantName"]

            self.after(0, self._update_connection_label)
            self.after(0, lambda: self._log(f"✅ Connected to: {self.tenant_name}"))
        except Exception as e:
            self.after(0, lambda: messagebox.showerror("Login failed", str(e)))

        self.after(0, lambda: self.conn_btn.configure(state="normal"))

    # ── Run refresh  ─────────────────────────────────────────────────────────
    def _run(self):
        path = self.workbook_path.get()
        if not path or not os.path.exists(path):
            messagebox.showerror("No file", "Please select your Excel file first.")
            return
        if not self.token or not self.tenant_id:
            messagebox.showerror("Not connected", "Please connect to Xero first.")
            return
        if os.path.isfile(path):
            try:
                open(path, "r+b").close()
            except PermissionError:
                messagebox.showerror("File in use",
                    "The Excel file is open. Please close it in Excel first.")
                return
        self.run_btn.configure(state="disabled")
        threading.Thread(target=self._do_refresh, daemon=True).start()

    def _do_refresh(self):
        path = self.workbook_path.get()
        try:
            wb = load_workbook(path)
        except Exception as e:
            self.after(0, lambda: messagebox.showerror("Cannot open file", str(e)))
            self.after(0, lambda: self.run_btn.configure(state="normal"))
            return

        periods, accounts = read_config(wb)
        self.after(0, lambda: self._log(
            f"Found {len(periods)} months and {len(accounts)} accounts."))

        token  = self.token["access_token"]
        tenant = self.tenant_id

        # Fetch P&L for each period
        all_data = {}
        for i, (col, year, month) in enumerate(periods):
            from_date = date(year, month, 1)
            to_date   = date(year, month, monthrange(year, month)[1])
            label     = from_date.strftime("%B %Y")
            self.after(0, lambda l=label: self._log(f"Fetching {l}…"))
            pct = int((i / len(periods)) * 90)
            self.after(0, lambda p=pct: self.progress.configure(value=p))

            url = f"{API_BASE}/Reports/ProfitAndLoss"
            url += "?" + urllib.parse.urlencode({
                "fromDate": from_date.strftime("%Y-%m-%d"),
                "toDate":   to_date.strftime("%Y-%m-%d"),
                "standardLayout":"true", "paymentsOnly":"false"})
            req = urllib.request.Request(url)
            req.add_header("Authorization",  f"Bearer {token}")
            req.add_header("Xero-Tenant-Id", tenant)
            req.add_header("Accept",         "application/json")
            try:
                with urllib.request.urlopen(req) as r:
                    data = json.loads(r.read())
                results = {}
                def walk(rows_list):
                    for row in rows_list:
                        rt = row.get("RowType","")
                        if rt in ("Section","SummaryRow"): walk(row.get("Rows",[]))
                        elif rt == "Row":
                            cells = row.get("Cells",[])
                            if len(cells) >= 2:
                                nm = cells[0].get("Value","").strip()
                                vl = cells[1].get("Value","0") or "0"
                                try: amt = float(vl.replace(",",""))
                                except: amt = 0.0
                                if nm: results[nm.lower()] = amt
                walk(data.get("Reports",[{}])[0].get("Rows",[]))
                all_data[(year,month)] = results
                self.after(0, lambda l=label, n=len(results):
                    self._log(f"  → {n} accounts for {l}"))
            except Exception as e:
                self.after(0, lambda l=label, e=e: self._log(f"  ⚠️ {l}: {e}"))
                all_data[(year,month)] = {}

        # Write back
        self.after(0, lambda: self._log("Writing to spreadsheet…"))
        self.after(0, lambda: self.progress.configure(value=95))
        ws = wb.active
        ws["A2"].value = self.tenant_name
        ws["A3"].value = f"Refreshed: {datetime.now().strftime('%d %b %Y %H:%M')}"
        written, not_found = 0, []

        for row, account_name in accounts:
            key = account_name.lower()
            for col, year, month in periods:
                period_data = all_data.get((year,month),{})
                cell = ws.cell(row=row, column=col)
                amount = None
                if key in period_data:
                    amount = period_data[key]
                else:
                    matches = [(k,v) for k,v in period_data.items() if key in k or k in key]
                    if len(matches) == 1: amount = matches[0][1]
                if amount is not None:
                    cell.value = amount
                    cell.number_format = NZD_FMT
                    cell.font  = Font(name="Arial", size=10)
                    cell.fill  = PatternFill("solid", fgColor="FFFFFF")
                    cell.alignment = Alignment(horizontal="right")
                    written += 1
                else:
                    cell.fill = PatternFill("solid", fgColor="FFFF00")
                    cell.value = None
                    if account_name not in not_found:
                        not_found.append(account_name)

        wb.save(path)
        self.after(0, lambda: self.progress.configure(value=100))
        self.after(0, lambda: self._log(f"✅ Done — {written} cells updated."))
        if not_found:
            self.after(0, lambda: self._log(
                f"⚠️ Not matched ({len(not_found)}): " + ", ".join(not_found)))
        self.after(0, lambda: messagebox.showinfo(
            "Complete", f"{written} cells updated.\nOpen {os.path.basename(path)} in Excel."))
        self.after(2000, lambda: self.progress.configure(value=0))
        self.after(0, lambda: self.run_btn.configure(state="normal"))

# ── Entry point  ─────────────────────────────────────────────────────────────
if __name__ == "__main__":
    App().mainloop()
